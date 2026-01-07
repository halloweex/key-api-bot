"""
Business logic services for sales reporting.

This module contains all business logic extracted from keycrm_api.py:
- Sales data aggregation
- Excel report generation
- TOP-10 product calculations
- Telegram file sending
"""
import logging
import time
import tempfile
import requests
import openpyxl

logger = logging.getLogger(__name__)

from openpyxl.styles import Font, PatternFill


class KeyCRMAPIError(Exception):
    """Custom exception for KeyCRM API errors."""
    def __init__(self, message: str, error_details: str = None):
        self.message = message
        self.error_details = error_details
        super().__init__(self.message)


class ReportGenerationError(Exception):
    """Custom exception for report generation errors."""
    def __init__(self, message: str, cause: Exception = None):
        self.message = message
        self.cause = cause
        super().__init__(self.message)
from datetime import datetime, timedelta
from collections import defaultdict
from typing import Dict, List, Tuple, Optional, Any
from zoneinfo import ZoneInfo

from bot.api_client import KeyCRMClient
from bot.config import (
    SOURCE_MAPPING,
    DEFAULT_TIMEZONE,
    TELEGRAM_MANAGER_IDS,
    RETURN_STATUS_IDS,
    API_PAGE_LIMIT,
    API_REQUEST_DELAY,
    ORDER_SYNC_BUFFER_HOURS
)


class ReportService:
    """Service for generating sales reports."""

    def __init__(self, api_client: KeyCRMClient):
        """
        Initialize report service.

        Args:
            api_client: KeyCRM API client instance
        """
        self.api = api_client

    def aggregate_sales_data(
        self,
        target_date: Any,
        tz_name: str = DEFAULT_TIMEZONE,
        exclude_status_id: Optional[int] = None,
        telegram_manager_ids: Optional[List[str]] = None
    ) -> Tuple[Dict, Dict, int, Dict, Dict]:
        """
        Get sales data aggregated by product and source for a specific date range.

        Args:
            target_date: The date(s) to get sales for:
                        - datetime.date object
                        - 'YYYY-MM-DD' string
                        - (start_date, end_date) tuple for a date range
            tz_name: Timezone name for local midnight-to-midnight (default: Europe/Kyiv)
            exclude_status_id: Status ID to exclude from results
            telegram_manager_ids: List of manager IDs to include for Telegram orders

        Returns:
            Tuple of:
            - sales_dict: {source_id: {product_name: total_qty, ...}, ...}
            - counts_dict: {source_id: total_order_count, ...}
            - total_orders: total orders retrieved
            - revenue_dict: {source_id: total_revenue, ...}
            - returns_dict: {source_id: {"count": N, "revenue": R}, ...}
        """
        # Use default manager IDs if not provided
        if telegram_manager_ids is None:
            telegram_manager_ids = TELEGRAM_MANAGER_IDS

        # Determine if we're dealing with a single date or date range
        if isinstance(target_date, tuple) and len(target_date) == 2:
            start_date, end_date = target_date
        else:
            # Single date - use the same date for start and end
            start_date = end_date = target_date

        # Normalize dates to string format if needed
        if not isinstance(start_date, str):
            start_date = start_date.strftime('%Y-%m-%d')
        if not isinstance(end_date, str):
            end_date = end_date.strftime('%Y-%m-%d')

        # Parse the strings into date objects
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        end = datetime.strptime(end_date, "%Y-%m-%d").date()

        # Initialize containers for orders
        all_orders = []
        returned_orders = []
        tz = ZoneInfo(tz_name)
        return_status_ids = set(RETURN_STATUS_IDS)

        # Calculate UTC boundaries for the ENTIRE period
        local_period_start = datetime(start.year, start.month, start.day, 0, 0, 0, tzinfo=tz)
        local_period_end = datetime(end.year, end.month, end.day, 23, 59, 59, tzinfo=tz)

        utc_period_start = local_period_start.astimezone(ZoneInfo("UTC"))
        utc_period_end = local_period_end.astimezone(ZoneInfo("UTC"))

        # Extend created_between to catch orders with delayed sync (see ORDER_SYNC_BUFFER_HOURS)
        utc_created_end = utc_period_end + timedelta(hours=ORDER_SYNC_BUFFER_HOURS)

        # Single request params for entire period
        params = {
            "include": "products,manager",
            "limit": API_PAGE_LIMIT,
            "filter[created_between]": f"{utc_period_start.strftime('%Y-%m-%d %H:%M:%S')}, {utc_created_end.strftime('%Y-%m-%d %H:%M:%S')}",
        }

        page = 1

        while True:
            params["page"] = page
            resp = self.api.get_orders(params)

            if isinstance(resp, dict) and resp.get("error"):
                error_msg = resp['error']
                logger.error(f"API Error while fetching orders: {error_msg}")
                raise KeyCRMAPIError("Failed to fetch orders from KeyCRM", error_msg)

            # Extract batch of orders
            batch = resp.get("data", [])
            if not batch:
                break

            # Process each order in the batch
            for order in batch:
                # Filter by ordered_at (not created_at) to match CRM UI
                ordered_at_str = order.get("ordered_at")
                if not ordered_at_str:
                    continue  # No ordered_at — skip
                ordered_at = datetime.fromisoformat(ordered_at_str.replace("Z", "+00:00"))
                if not (utc_period_start <= ordered_at <= utc_period_end):
                    continue  # Outside target period

                # Get manager ID if present
                manager_id = None
                if "manager" in order and order["manager"]:
                    manager = order["manager"]
                    manager_id = manager.get("id")

                # Get the status ID
                status_id = None
                if "status_id" in order:
                    status_id = order["status_id"]
                elif "status" in order and order["status"] and "id" in order["status"]:
                    status_id = order["status"]["id"]

                # Convert to string for comparison
                status_id_str = str(status_id) if status_id is not None else "None"
                manager_id_str = str(manager_id) if manager_id is not None else "None"
                exclude_id_str = str(exclude_status_id) if exclude_status_id is not None else None

                source_id = order.get("source_id")

                # Track returned orders separately
                if status_id in return_status_ids:
                    returned_orders.append(order)
                    continue

                if exclude_id_str and status_id_str == exclude_id_str:
                    continue

                # Handle Telegram orders separately
                if source_id == 2:  # Telegram source ID
                    # For Telegram, only include orders from specified managers
                    if telegram_manager_ids and manager_id_str not in telegram_manager_ids:
                        continue

                all_orders.append(order)

            # Or if we got fewer results than requested
            if len(batch) < params["limit"]:
                break

            page += 1

            # Add a small delay to avoid hitting API rate limits
            time.sleep(API_REQUEST_DELAY)

        # Aggregate sales data
        sales = defaultdict(lambda: defaultdict(int))
        revenue = defaultdict(float)

        for order in all_orders:
            src = order.get("source_id") or "unknown"
            for prod in order.get("products", []):
                name = prod.get("name") or f"#{prod.get('id')}"
                qty = int(prod.get("quantity", 0))
                sales[src][name] += qty
            order_total = float(order.get("grand_total", 0))
            revenue[src] += order_total

        # Count orders per-source
        counts = defaultdict(int)
        for order in all_orders:
            src = order.get("source_id") or "unknown"
            counts[src] += 1

        # Process returns data
        returns_data = defaultdict(lambda: {"count": 0, "revenue": 0})
        for order in returned_orders:
            src = order.get("source_id") or "unknown"
            returns_data[src]["count"] += 1
            order_total = float(order.get("grand_total", 0))
            returns_data[src]["revenue"] += order_total

        # Convert to plain dicts
        sales_dict = {src: dict(prod_map) for src, prod_map in sales.items()}
        counts_dict = dict(counts)
        total_orders = len(all_orders)
        revenue_dict = dict(revenue)
        returns_dict = {src: dict(data) for src, data in returns_data.items()}

        return sales_dict, counts_dict, total_orders, revenue_dict, returns_dict

    def calculate_top10_products(
        self,
        target_date: Any,
        source_id: int,
        limit: int = 10,
        tz_name: str = DEFAULT_TIMEZONE,
        exclude_status_id: Optional[int] = None
    ) -> Tuple[List[Tuple[str, int, float]], int]:
        """
        Get TOP products for a specific source.

        Args:
            target_date: Date or date range
            source_id: Source ID (1=Instagram, 2=Telegram, 4=Shopify)
            limit: Number of top products to return (default: 10)
            tz_name: Timezone name
            exclude_status_id: Status ID to exclude

        Returns:
            Tuple of:
            - List of (product_name, quantity, percentage) tuples
            - Total quantity across all products
        """
        sales_dict, _, _, _, _ = self.aggregate_sales_data(
            target_date=target_date,
            tz_name=tz_name,
            exclude_status_id=exclude_status_id
        )

        # Get products for specific source
        products = sales_dict.get(source_id, {})
        total_quantity = sum(products.values())

        # Sort by quantity and get top N
        sorted_products = sorted(products.items(), key=lambda x: x[1], reverse=True)
        top_products = sorted_products[:limit]

        # Calculate percentages
        result = []
        for product_name, quantity in top_products:
            percentage = (quantity / total_quantity * 100) if total_quantity > 0 else 0
            result.append((product_name, quantity, percentage))

        return result, total_quantity

    def generate_excel_report(
        self,
        target_date: Any,
        bot_token: str,
        chat_id: int,
        tz_name: str = DEFAULT_TIMEZONE,
        exclude_status_id: Optional[int] = None,
        telegram_manager_ids: Optional[List[str]] = None
    ) -> bool:
        """
        Generate a sales summary Excel file and send it to a Telegram chat.

        Args:
            target_date: The date(s) to get sales for
            bot_token: Telegram bot API token
            chat_id: Telegram chat ID to send the file to
            tz_name: Timezone name (default: Europe/Kyiv)
            exclude_status_id: Status ID to exclude from results
            telegram_manager_ids: List of manager IDs to include for Telegram orders

        Returns:
            True if the file was sent successfully, False otherwise
        """
        try:
            # Get sales data
            sales_data, counts, total, revenue_data, returns_data = self.aggregate_sales_data(
                target_date=target_date,
                tz_name=tz_name,
                exclude_status_id=exclude_status_id,
                telegram_manager_ids=telegram_manager_ids
            )

            # Determine the date range string
            if isinstance(target_date, tuple) and len(target_date) == 2:
                date_range_str = f"{target_date[0]}_to_{target_date[1]}"
                display_date = f"{target_date[0]} to {target_date[1]}"
            else:
                if isinstance(target_date, str):
                    date_range_str = target_date
                    display_date = target_date
                else:
                    date_range_str = target_date.strftime('%Y-%m-%d')
                    display_date = date_range_str

            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sales Summary"

            # Define styles
            header_font = Font(name='Arial', size=12, bold=True)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            source_font = Font(name='Arial', size=11, bold=True)
            source_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

            # Write header rows
            ws['A1'] = f"Sales Summary for {display_date} (Timezone: {tz_name})"
            ws['A1'].font = Font(name='Arial', size=14, bold=True)
            ws.merge_cells('A1:C1')

            ws['A2'] = f"Total Orders: {total}"
            ws['A2'].font = header_font

            row = 3
            if exclude_status_id is not None:
                ws[f'A{row}'] = f"Excluded Orders with Status ID: {exclude_status_id}"
                ws[f'A{row}'].font = Font(name='Arial', size=10)
                row += 1

            if telegram_manager_ids:
                ws[f'A{row}'] = f"Filtered Telegram Orders to Managers: {', '.join(str(id) for id in telegram_manager_ids)}"
                ws[f'A{row}'].font = Font(name='Arial', size=10)
                row += 1

            row += 1  # Empty row

            # For each source
            for src_id, products in sales_data.items():
                # Map source ID to name
                try:
                    src_key = int(src_id)
                except (ValueError, TypeError):
                    src_key = src_id

                src_name = SOURCE_MAPPING.get(src_key, src_id)
                order_count = counts.get(src_id, 0)

                # Write source name
                ws[f'A{row}'] = f"Source: {src_name}"
                ws[f'A{row}'].font = source_font
                ws[f'A{row}'].fill = source_fill
                ws.merge_cells(f'A{row}:C{row}')
                row += 1

                ws[f'A{row}'] = f"Total Orders: {order_count}"
                ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
                row += 1

                # Average check
                total_revenue = revenue_data.get(src_id, 0)
                avg_check = total_revenue / order_count if order_count > 0 else 0

                ws[f'A{row}'] = f"Average Check: {avg_check:.2f} UAH"
                ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
                ws[f'A{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                row += 1

                # Column headers
                ws[f'A{row}'] = "Product"
                ws[f'B{row}'] = "Quantity"
                ws[f'A{row}'].font = header_font
                ws[f'B{row}'].font = header_font
                ws[f'A{row}'].fill = header_fill
                ws[f'B{row}'].fill = header_fill
                row += 1

                # Sort products by quantity
                sorted_products = sorted(products.items(), key=lambda x: x[1], reverse=True)

                # Write products
                for product_name, quantity in sorted_products:
                    ws[f'A{row}'] = product_name
                    ws[f'B{row}'] = quantity
                    row += 1

                # Empty row
                row += 1

            # Set column widths
            ws.column_dimensions['A'].width = 60
            ws.column_dimensions['B'].width = 15

            # Create temporary file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_file_path = temp_file.name

            # Save Excel file
            wb.save(temp_file_path)

            # Create filename
            current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"sales_report_{date_range_str}_{current_time}.xlsx"

            # Send to Telegram
            url = f"https://api.telegram.org/bot{bot_token}/sendDocument"

            with open(temp_file_path, 'rb') as file:
                files = {'document': (filename, file)}
                data = {'chat_id': chat_id, 'caption': f"Sales Report for {display_date}"}
                response = requests.post(url, data=data, files=files)

            if response.status_code == 200:
                logger.info(f"Sales report for {display_date} sent to Telegram successfully")
                return True
            else:
                logger.error(f"Failed to send report to Telegram: {response.status_code} - {response.text}")
                return False

        except Exception as e:
            logger.error(f"Error generating or sending sales report: {str(e)}", exc_info=True)
            return False
