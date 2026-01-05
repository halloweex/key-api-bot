import os
import requests
import json
from datetime import datetime, time, timedelta
from collections import defaultdict
from dotenv import load_dotenv
import time
import tempfile
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from zoneinfo import ZoneInfo

load_dotenv()

KEYCRM_API_KEY = os.getenv("KEYCRM_API_KEY")

# Source dictionary mapping source_id to source name
source_dct = {1: 'Instagram', 2: 'Telegram', 3: 'Opencart', 4: 'Shopify'}

class KeyCRMAPI:
    """
    A Python client for the KeyCRM OpenAPI.
    """

    def __init__(self, api_key, base_url="https://openapi.keycrm.app/v1"):
        """
        Initialize the KeyCRM API client.

        Args:
            api_key (str): Your KeyCRM API key
            base_url (str): The base URL for the KeyCRM API
        """
        print(f"Initializing KeyCRM with API key: {api_key[:10] if api_key else 'NONE'}...")
        self.api_key = api_key
        self.base_url = base_url
        self.headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json"
        }

    def _make_request(self, method, endpoint, params=None, data=None):
        """
        Make a request to the KeyCRM API.

        Args:
            method (str): HTTP method (GET, POST, PUT, DELETE)
            endpoint (str): API endpoint
            params (dict, optional): Query parameters
            data (dict, optional): Data to send in the request body

        Returns:
            dict: API response
        """
        url = f"{self.base_url}/{endpoint}"

        try:
            if method.upper() == "GET":
                response = requests.get(url, headers=self.headers, params=params)
            elif method.upper() == "POST":
                response = requests.post(url, headers=self.headers, json=data)
            elif method.upper() == "PUT":
                response = requests.put(url, headers=self.headers, json=data)
            elif method.upper() == "DELETE":
                response = requests.delete(url, headers=self.headers, params=params)
            else:
                raise ValueError(f"Unsupported HTTP method: {method}")

            response.raise_for_status()  # Raise an exception for 4XX/5XX responses

            if response.content:
                return response.json()
            return {"status": "success"}

        except requests.exceptions.RequestException as e:
            print(f"Error making request: {e}")
            if hasattr(e, 'response') and e.response is not None:
                print(f"Response status code: {e.response.status_code}")
                print(f"Response body: {e.response.text}")
            return {"error": str(e)}

    # Orders
    def get_orders(self, params=None):
        """Get a list of orders with optional filtering"""
        return self._make_request("GET", "order", params=params)

    def get_order(self, order_id, include_products=True):
        """
        Get a specific order by ID

        Args:
            order_id: The ID of the order to retrieve
            include_products: Whether to include products information (default True)
        """
        params = {"include": "products,buyer"} if include_products else None
        return self._make_request("GET", f"order/{order_id}", params=params)

    def create_order(self, order_data):
        """Create a new order"""
        return self._make_request("POST", "order", data=order_data)

    def update_order(self, order_id, order_data):
        """Update an existing order"""
        return self._make_request("PUT", f"order/{order_id}", data=order_data)

    # Customers
    def get_customers(self, params=None):
        """Get a list of customers with optional filtering"""
        return self._make_request("GET", "customer", params=params)

    def get_customer(self, customer_id):
        """Get a specific customer by ID"""
        return self._make_request("GET", f"customer/{customer_id}")

    def create_customer(self, customer_data):
        """Create a new customer"""
        return self._make_request("POST", "customer", data=customer_data)

    def update_customer(self, customer_id, customer_data):
        """Update an existing customer"""
        return self._make_request("PUT", f"customer/{customer_id}", data=customer_data)

    # Products
    def get_products(self, params=None):
        """Get a list of products with optional filtering"""
        return self._make_request("GET", "product", params=params)

    def get_product(self, product_id):
        """Get a specific product by ID"""
        return self._make_request("GET", f"product/{product_id}")

    # Statuses
    def get_statuses(self):
        """Get all available order statuses"""
        return self._make_request("GET", "status")

    def get_top_products_site_instagram(
            self,
            target_date,
            limit=10,
            tz_name="Europe/Kyiv",
            exclude_status_id=None):
        """Get TOP products for Site (Opencart) + Instagram combined."""

        # Use the updated get_sales_by_product_and_source_for_date that returns 4 values
        sales_dict, _, _, _, _ = self.get_sales_by_product_and_source_for_date(
            target_date=target_date,
            tz_name=tz_name,
            exclude_status_id=exclude_status_id
        )

        # Combine Site (Opencart id=3) and Instagram (id=1) products
        combined_products = defaultdict(int)
        total_quantity = 0

        for source_id in [1, 3]:  # Instagram and Opencart
            if source_id in sales_dict:
                for product_name, quantity in sales_dict[source_id].items():
                    combined_products[product_name] += quantity
                    total_quantity += quantity

        # Sort by quantity and get top N
        sorted_products = sorted(combined_products.items(), key=lambda x: x[1], reverse=True)
        top_products = sorted_products[:limit]

        # Calculate percentages
        result = []
        for product_name, quantity in top_products:
            percentage = (quantity / total_quantity * 100) if total_quantity > 0 else 0
            result.append((product_name, quantity, percentage))

        return result, total_quantity

    def get_sales_by_product_and_source_for_date(
            self,
            target_date,
            tz_name="Europe/Kyiv",
            exclude_status_id=None,
            telegram_manager_ids=None):
        """
        Get sales data aggregated by product and source for a specific local-date window
        or date range, using KeyCRM's `created_between` filter.

        Args:
            target_date (date or str or tuple): The date(s) to get sales for:
                                              - datetime.date object
                                              - 'YYYY-MM-DD' string
                                              - (start_date, end_date) tuple for a date range
            tz_name (str): Timezone name to define your local midnight-to-midnight.
                           Default is "Europe/Kyiv" (Kyiv timezone).
            exclude_status_id (int, optional): Status ID to exclude from results
            telegram_manager_ids (list, optional): List of manager IDs to include for Telegram orders

        Returns:
            tuple:
                sales_dict (dict): { source_id: { product_name: total_qty, … }, … }
                counts_dict (dict): { source_id: total_order_count, … }
                total_orders (int): total orders retrieved that passed all filters
        """
        # Ensure telegram_manager_ids is a list of strings for easier comparison
        if telegram_manager_ids is None:
            telegram_manager_ids = []
        telegram_manager_ids = [str(id) for id in telegram_manager_ids]

        # Determine if we're dealing with a single date or date range
        if isinstance(target_date, tuple) and len(target_date) == 2:
            start_date, end_date = target_date
        else:
            # Single date - use the same date for start and end
            start_date = end_date = target_date

        # Normalize dates to string format if needed
        if not isinstance(start_date, str):
            start_date = start_date.strftime('%Y-%m-%d')
        if not isinstance(end_date, str):
            end_date = end_date.strftime('%Y-%m-%d')

        # Parse the strings into date objects
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        end = datetime.strptime(end_date, "%Y-%m-%d").date()

        # Initialize containers for orders
        all_orders = []
        excluded_orders = []
        filtered_telegram_orders = []
        returned_orders = []
        current = start
        print(f"DEBUG tz_name: {tz_name}")
        tz = ZoneInfo(tz_name)
        #tz = ZoneInfo("Europe/Kyiv")  # Использует актуальную системную tzdata
        # For each day in our date range
        while current <= end:
            # Create local midnight and end of day
            local_start = datetime(current.year, current.month, current.day, 0, 0, 0, tzinfo=tz)
            local_end = datetime(current.year, current.month, current.day, 23, 59, 59, tzinfo=tz)

            utc_start = local_start.astimezone(ZoneInfo("UTC"))
            utc_end = local_end.astimezone(ZoneInfo("UTC"))

            utc_start_str = utc_start.strftime("%Y-%m-%d %H:%M:%S")
            utc_end_str = utc_end.strftime("%Y-%m-%d %H:%M:%S")

            # For each source, get the full order details
            params = {
                "include": "products,manager",  # Include manager for filtering
                "limit": 50,  # KeyCRM API appears to have a 50-item limit per page
                "filter[created_between]": f"{utc_start_str}, {utc_end_str}",
            }

            # Get all orders for this day
            day_orders = []
            page = 1
            total_pages = None

            while True:
                params["page"] = page
                resp = self.get_orders(params)

                if isinstance(resp, dict) and resp.get("error"):
                    print(f"API Error: {resp['error']}")
                    break

                # Extract batch of orders
                batch = resp.get("data", [])
                if not batch:
                    break

                # Try to extract pagination metadata if available
                if total_pages is None and resp.get("meta") and "last_page" in resp.get("meta", {}):
                    total_pages = resp["meta"]["last_page"]
                    total_items = resp["meta"].get("total", "unknown")

                # Process each order in the batch
                for order in batch:
                    # Get manager ID if present
                    manager_id = None
                    if "manager" in order and order["manager"]:
                        manager = order["manager"]
                        manager_id = manager.get("id")

                    # Get the status ID
                    status_id = None
                    if "status_id" in order:
                        status_id = order["status_id"]
                    elif "status" in order and order["status"] and "id" in order["status"]:
                        status_id = order["status"]["id"]

                    # Convert to string for comparison
                    status_id_str = str(status_id) if status_id is not None else "None"
                    manager_id_str = str(manager_id) if manager_id is not None else "None"
                    exclude_id_str = str(exclude_status_id) if exclude_status_id is not None else None

                    source_id = order.get("source_id")

                    # Track returned orders separately (status_id 22)
                    if status_id in [22, 19]:
                        returned_orders.append(order)
                        continue

                    if exclude_id_str and status_id_str == exclude_id_str:
                        excluded_orders.append(order)
                        continue

                    # Handle Telegram orders separately
                    if source_id == 2:  # Telegram source ID
                        # For Telegram, only include orders from specified managers
                        if telegram_manager_ids and manager_id_str not in telegram_manager_ids:
                            filtered_telegram_orders.append(order)
                            continue

                    # Add order to filtered list
                    day_orders.append(order)

                # Check if we've reached the last page
                if total_pages and page >= total_pages:
                    break

                # Or if we got fewer results than requested
                if len(batch) < params["limit"]:
                    break

                page += 1

                # Add a small delay to avoid hitting API rate limits
                time.sleep(0.5)

            # Add this day's orders to the overall list
            all_orders.extend(day_orders)

            # Move to next day
            current += timedelta(days=1)

        # Once we have all filtered orders for the entire date range, perform aggregation
        sales = defaultdict(lambda: defaultdict(int))
        revenue = defaultdict(float)
        for order in all_orders:
            src = order.get("source_id") or "unknown"
            for prod in order.get("products", []):
                name = prod.get("name") or f"#{prod.get('id')}"
                qty = int(prod.get("quantity", 0))
                sales[src][name] += qty
            order_total = float(order.get("grand_total", 0))
            revenue[src] += order_total

        # Count orders per-source
        counts = defaultdict(int)
        for order in all_orders:
            src = order.get("source_id") or "unknown"
            counts[src] += 1

        # Return plain dicts + total (exactly 3 values as expected)
        sales_dict = {src: dict(prod_map) for src, prod_map in sales.items()}
        counts_dict = dict(counts)
        total_orders = len(all_orders)
        revenue_dict = dict(revenue)

        # Process returns data
        returns_data = defaultdict(lambda: {"count": 0, "revenue": 0})
        for order in returned_orders:
            src = order.get("source_id") or "unknown"
            returns_data[src]["count"] += 1
            order_total = float(order.get("grand_total", 0))
            returns_data[src]["revenue"] += order_total

        returns_dict = {src: dict(data) for src, data in returns_data.items()}
        return sales_dict, counts_dict, total_orders, revenue_dict, returns_dict

    def send_sales_summary_excel_to_telegram(
            self,
            target_date,
            bot_token,
            chat_id,
            tz_name="Europe/Kyiv",
            exclude_status_id=None,
            telegram_manager_ids=None):
        """
        Generate a sales summary Excel file and send it to a Telegram chat.

        Args:
            target_date (date or str or tuple): The date(s) to get sales for
            bot_token (str): Telegram bot API token
            chat_id (str or int): Telegram chat ID to send the file to
            tz_name (str): Timezone name (default: "Europe/Kyiv")
            exclude_status_id (int, optional): Status ID to exclude from results
            telegram_manager_ids (list, optional): List of manager IDs to include for Telegram orders

        Returns:
            bool: True if the file was sent successfully, False otherwise
        """
        try:
            # Get sales data
            sales_data, counts, total, revenue_data, returns_data = self.get_sales_by_product_and_source_for_date(
                target_date=target_date,
                tz_name=tz_name,
                exclude_status_id=exclude_status_id,
                telegram_manager_ids=telegram_manager_ids
            )

            # Determine the date range string for the report header and filename
            if isinstance(target_date, tuple) and len(target_date) == 2:
                date_range_str = f"{target_date[0]}_to_{target_date[1]}"
                display_date = f"{target_date[0]} to {target_date[1]}"
            else:
                if isinstance(target_date, str):
                    date_range_str = target_date
                    display_date = target_date
                else:
                    date_range_str = target_date.strftime('%Y-%m-%d')
                    display_date = date_range_str

            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sales Summary"

            # Define styles
            header_font = Font(name='Arial', size=12, bold=True)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            source_font = Font(name='Arial', size=11, bold=True)
            source_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

            # Write header rows
            ws['A1'] = f"Sales Summary for {display_date} (Timezone: {tz_name})"
            ws['A1'].font = Font(name='Arial', size=14, bold=True)
            ws.merge_cells('A1:C1')

            ws['A2'] = f"Total Orders: {total}"
            ws['A2'].font = header_font

            row = 3
            if exclude_status_id is not None:
                ws[f'A{row}'] = f"Excluded Orders with Status ID: {exclude_status_id}"
                ws[f'A{row}'].font = Font(name='Arial', size=10)
                row += 1

            if telegram_manager_ids:
                ws[
                    f'A{row}'] = f"Filtered Telegram Orders to Managers: {', '.join(str(id) for id in telegram_manager_ids)}"
                ws[f'A{row}'].font = Font(name='Arial', size=10)
                row += 1

            row += 1  # Empty row as separator

            # For each source
            for src_id, products in sales_data.items():
                # Map numeric/string src_id → human name if possible
                try:
                    src_key = int(src_id)
                except (ValueError, TypeError):
                    src_key = src_id

                src_name = source_dct.get(src_key, src_id)
                order_count = counts.get(src_id, 0)

                # Write source name and order count
                ws[f'A{row}'] = f"Source: {src_name}"
                ws[f'A{row}'].font = source_font
                ws[f'A{row}'].fill = source_fill
                ws.merge_cells(f'A{row}:C{row}')
                row += 1

                ws[f'A{row}'] = f"Total Orders: {order_count}"
                ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
                row += 1

                # ADD these lines:
                total_revenue = revenue_data.get(src_id, 0)
                avg_check = total_revenue / order_count if order_count > 0 else 0

                ws[f'A{row}'] = f"Average Check: {avg_check:.2f} UAH"
                ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
                ws[f'A{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                row += 1

                # Column headers
                ws[f'A{row}'] = "Product"
                ws[f'B{row}'] = "Quantity"
                ws[f'A{row}'].font = header_font
                ws[f'B{row}'].font = header_font
                ws[f'A{row}'].fill = header_fill
                ws[f'B{row}'].fill = header_fill
                row += 1

                # Sort products by quantity in descending order
                sorted_products = sorted(products.items(), key=lambda x: x[1], reverse=True)

                # Write each product and its quantity
                for product_name, quantity in sorted_products:
                    ws[f'A{row}'] = product_name
                    ws[f'B{row}'] = quantity
                    row += 1

                # Add empty row between sources
                row += 1

            # Set column widths
            ws.column_dimensions['A'].width = 60
            ws.column_dimensions['B'].width = 15

            # Create a temporary file to save the Excel
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_file_path = temp_file.name

            # Save Excel file
            wb.save(temp_file_path)

            # Create a human-readable filename
            current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"sales_report_{date_range_str}_{current_time}.xlsx"

            # Send the file to Telegram
            url = f"https://api.telegram.org/bot{bot_token}/sendDocument"

            with open(temp_file_path, 'rb') as file:
                files = {'document': (filename, file)}
                data = {'chat_id': chat_id, 'caption': f"Sales Report for {display_date}"}
                response = requests.post(url, data=data, files=files)

            # Print a success message with summary to console
            if response.status_code == 200:
                print(f"\nSales report for {display_date} sent to Telegram successfully.")
                return True
            else:
                print(f"Failed to send report to Telegram. Error: {response.text}")
                return False

        except Exception as e:
            print(f"Error generating or sending sales report: {str(e)}")
            return False

    def get_top_products_by_source(
            self,
            target_date,
            source_id,
            limit=10,
            tz_name="Europe/Kyiv",
            exclude_status_id=None):
        """Get TOP products for a specific source."""

        sales_dict, _, _, _, _ = self.get_sales_by_product_and_source_for_date(
            target_date=target_date,
            tz_name=tz_name,
            exclude_status_id=exclude_status_id
        )

        # Get products for specific source
        products = sales_dict.get(source_id, {})
        total_quantity = sum(products.values())

        # Sort by quantity and get top N
        sorted_products = sorted(products.items(), key=lambda x: x[1], reverse=True)
        top_products = sorted_products[:limit]

        # Calculate percentages
        result = []
        for product_name, quantity in top_products:
            percentage = (quantity / total_quantity * 100) if total_quantity > 0 else 0
            result.append((product_name, quantity, percentage))

        return result, total_quantity